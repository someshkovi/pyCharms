import paramiko
import sys
from paramiko import SSHClient
import json
import pandas as pd
import os
import glob
import xlrd
from openpyxl import load_workbook
from datetime import date
from datetime import datetime

hostname1 = '10.66.47.11'
hostname2 = '10.66.47.12'
hostname = hostname2
port = 22
username = 'root'
password = r'password'

Date = datetime.now().strftime("%d-%m-%Y")

def ispriamry(hostname):
    stdin, stdout, stderr = client.exec_command(r'. /usr/adic/.profile ;snhamgr status')
    status = str(stdout.read(), 'utf-8')
    print("snhamgr status:")
    print(status)
    if status.split("\n")[1] == 'LocalStatus=primary':
        print(hostname +" is primary")
        return(1)
    else:
        retrun(0)

cin = r'. /usr/adic/.profile ;'
cmd0 = 'snhamgr status'
cmd1 = 'fsstate'
cmd2 = 'snbackup -s'

cmd3 = 'df -k | grep -E '+'\'stornext|shared\''
cmd4 = 'tail -n 10 /usr/adic/TSM/logs/trace/trace_01'
cmd5 = 'tail -n 10 /usr/adic/TSM/logs/tac/tac_00'
cmd6 = 'tail -n 10 /var/log/messages'

#path = os.getcwd()
path = r'C:\Users\somesh\Desktop\util\data'
path = path+'\\'+str(Date)
try:
    os.mkdir(path)
except FileExistsError:
    print(FileExistsError)
os.chdir(path)

try:
    client = paramiko.SSHClient()
    client.load_system_host_keys()
    client.set_missing_host_key_policy(paramiko.WarningPolicy)

    client.connect(hostname, port=port, username=username, password=password)

    stdin, stdout, stderr = client.exec_command(cin+cmd0)
    status = str(stdout.read(), 'utf-8')
    #print("snhamgr status:")
    #print(status)

    if status.split("\n")[1] == 'LocalStatus=primary':
        print(hostname +" is primary")
    else:
        hostname = hostname1
        client.connect(hostname, port=port, username=username, password=password)

    stdin, stdout, stderr = client.exec_command(cmd3)
    disk = str(stdout.read(), 'utf-8')
    print("Disk Usage:")
    print(disk)

    stdin, stdout, stderr = client.exec_command(cin+"vsmedqry -a |grep -i 'media id' |awk '{print $3}'")
    sys.stdout = open('medlist.csv', 'w')
    fsmedlist = str(stdout.read(), 'utf-8')
    print(fsmedlist)
    sys.stdout = open('temp.txt', 'w')
    medlist = pd.read_csv("medlist.csv")
    for f in range(len(medlist)):
        mediaid = medlist.iloc[f][0]
        cm = 'fsmedinfo '+mediaid+' -F JSON'
        stdin, stdout, stderr = client.exec_command(cin+cm)
        sys.stdout = open(mediaid+'.json', 'w')
        fsmedinfo = str(stdout.read(), 'utf-8')
        print(fsmedinfo)
    sys.stdout = open('temp.txt', 'w')

finally:
    client.close()

list_of_files = [f for f in glob.glob("*.json")]
i = 0
for f in list_of_files:
    try:
        if i ==0:
            data = json.load(open(f))
            mdusage = pd.DataFrame(data["medias"])
            i = 1
        else:
            data = json.load(open(f))
            Media = pd.DataFrame(data["medias"])
            mdusage = pd.concat([mdusage,Media],sort=False)
    except:
        print(f)


mdusage['percentUsed'] = mdusage.percentUsed.apply(lambda c: float(c))
dfpp = mdusage.groupby('classId', as_index=False).agg({"percentUsed": "mean"})
dfpc = mdusage.groupby('classId', as_index=False).agg({"mediaId": "count"})
dfArc = mdusage[mdusage['currentArchive'] == 'Archive'].groupby('classId', as_index=False).agg({"mediaId": "count"})
dfArc = dfArc.rename(columns={"mediaId":"Archive"})
dfAva = mdusage[mdusage['currentArchive'] == 'Archive']
dfAva = dfAva[dfAva['mediaStatus'] == 'UNAVAIL'].groupby('classId', as_index=False).agg({"mediaId": "count"})
dfAva = dfAva.rename(columns={"mediaId":"UNAVAIL"})
result = dfpp.merge(dfpc, how = 'left', on='classId')
result = result.merge(dfArc, how = 'left', on='classId')
result = result.merge(dfAva, how = 'left', on='classId')
result['percentUsed'] = result.apply(lambda percentUsed : round((result['percentUsed']*result['mediaId'])/(mdusage['mediaId'].count()),2))

result['Date'] = result.apply(lambda row: Date, axis =1)


spath = r'C:\Users\somesh\Desktop\util'
os.chdir(spath)


writer = pd.ExcelWriter('PolicySummary.xlsx', engine='openpyxl')

"""
#for first time
result.to_excel(writer, sheet_name='Sheet1')
writer.save()
"""

# try to open an existing workbook
writer.book = load_workbook('PolicySummary.xlsx')
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(r'PolicySummary.xlsx')
# write out the new sheet
result.to_excel(writer,index=True,header=False,startrow=len(reader)+1)
writer.close()



#Convertiting data to df
k = disk.split('\n')
try:
    k.remove('')
except ValueError as error:
    pass
for i in range(0,len(k)):
    k[i] = k[i].split(' ')
    R = len(k[i])
    for r in range(0,R):
        try:
            k[i].remove('')
        except ValueError as error:
            pass

diskusage = pd.DataFrame(k, columns = ['Filesystem', 'Size', 'Used', 'Avail', 'UseP', 'Mount'])
diskusage['Size'] = diskusage['Size'].astype('float')
diskusage['Used'] = diskusage['Used'].astype('float')
diskusage['Avail'] = diskusage['Avail'].astype('float')
diskusage['UseP'] = diskusage['UseP'].str.rstrip('%').astype('float')
diskusage['Size_TB'] = diskusage.Size.apply(lambda c: c/(1024*1024*1024))
diskusage['Size_TB'] = diskusage.Size_TB.apply(lambda c: round(c))
diskusage['Date'] = diskusage.apply(lambda row: Date, axis =1)
diskusageS = diskusage[['Filesystem','Size_TB','UseP','Date']]
diskusageS = diskusageS[diskusageS.Filesystem != 'Shared_HA_CX1652CAB00009']
TMDCSpace = round(diskusage['Size'].sum()/(1024*1024*1024))
TMDCSpaceUse = diskusage['Used'].sum()/(1024*1024*1024)
TMDCSpaceUseP = round(TMDCSpaceUse*100/TMDCSpace,2)
diskusageS = diskusageS.append({'Filesystem' : 'SNFStotal' , 'Size_TB' : TMDCSpace, 'UseP' : TMDCSpaceUseP, 'Date':Date} , ignore_index=True)

avgmdusgP = round(mdusage['percentUsed'].mean(),2)
avgmdsizeTB = round(5980033680870/(1024*1024*1024*1024),2)
totalmdsizeTB = round(mdusage['mediaId'].count()*avgmdsizeTB,2)
Summary = diskusageS.append({'Filesystem' : 'TapeLibrary' , 'Size_TB' : totalmdsizeTB, 'UseP' : avgmdusgP, 'Date':Date} , ignore_index=True)

writer = pd.ExcelWriter('Summary.xlsx', engine='openpyxl')

"""
#for first time
Summary.to_excel(writer, sheet_name='Sheet1')
writer.save()
"""

# try to open an existing workbook
writer.book = load_workbook('Summary.xlsx')
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(r'Summary.xlsx')
# write out the new sheet
Summary.to_excel(writer,index=True,header=False,startrow=len(reader)+1)
writer.close()
