import paramiko
import sys
from paramiko import SSHClient
import json
import pandas as pd
import os
import glob
import xlrd
from openpyxl import load_workbook
from datetime import date
from datetime import datetime

from storageUtilUploadDb import *

result,Summary = getStorageUtil()
spath = r'C:\Users\somesh\Desktop\util'
os.chdir(spath)

writer = pd.ExcelWriter('PolicySummary.xlsx', engine='openpyxl')

"""
#for first time
result.to_excel(writer, sheet_name='Sheet1')
writer.save()
"""

# try to open an existing workbook
writer.book = load_workbook('PolicySummary.xlsx')
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(r'PolicySummary.xlsx')
# write out the new sheet
result.to_excel(writer,index=True,header=False,startrow=len(reader)+1)
writer.close()



writer = pd.ExcelWriter('Summary.xlsx', engine='openpyxl')

"""
#for first time
Summary.to_excel(writer, sheet_name='Sheet1')
writer.save()
"""

# try to open an existing workbook
writer.book = load_workbook('Summary.xlsx')
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(r'Summary.xlsx')
# write out the new sheet
Summary.to_excel(writer,index=True,header=False,startrow=len(reader)+1)
writer.close()
